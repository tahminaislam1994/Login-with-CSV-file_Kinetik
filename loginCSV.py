from selenium import webdriver

import openpyxl as O

driver = webdriver.Chrome(executable_path="C:\Driver\chromedriver\chromedriver.exe")

driver.get("https://shopnshipbd.com/login")
driver.maximize_window()

Excel_file = "E:\logintest.xlsx"
Excel_worksheet = "Data1"
wb = O.load_workbook(Excel_file)
ws = wb[Excel_worksheet]
row_num = ws.max_row
col_num = ws.max_column

print("The no. of row is",row_num,"and the number of column is",col_num)
row = 2

driver.find_element_by_xpath("//input[@id='emailOrPhone']").send_keys((ws.cell(row,1).value))
driver.find_element_by_xpath("//input[@id='password']").send_keys((ws.cell(row,2).value))

#driver.find_element_by_xpath("//input[@id='emailOrPhone']").send_keys((ws.cell(row+1,1).value))
#driver.find_element_by_xpath("//input[@id='password']").send_keys((ws.cell(row+1,2).value))

driver.find_element_by_xpath("//button[normalize-space()='Sign in']").click()
